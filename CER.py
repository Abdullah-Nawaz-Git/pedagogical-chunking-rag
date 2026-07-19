import openpyxl
import re

def normalize(s):
    """Minimal normalization: collapse newlines/whitespace, strip ends."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace('\n', ' ').replace('\r', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def edit_distance(a, b):
    """Standard Levenshtein edit distance via DP (character-level)."""
    n, m = len(a), len(b)
    if n == 0:
        return m
    if m == 0:
        return n
    prev = list(range(m + 1))
    for i in range(1, n + 1):
        curr = [i] + [0] * m
        ai = a[i - 1]
        for j in range(1, m + 1):
            cost = 0 if ai == b[j - 1] else 1
            curr[j] = min(
                prev[j] + 1,       # deletion
                curr[j - 1] + 1,   # insertion
                prev[j - 1] + cost # substitution
            )
        prev = curr
    return prev[m]

def cer(hyp, ref):
    """CER = edit_distance(hyp, ref) / len(ref), on normalized text."""
    ref_n = normalize(ref)
    hyp_n = normalize(hyp)
    if len(ref_n) == 0:
        return None
    d = edit_distance(hyp_n, ref_n)
    return d / len(ref_n), d, len(ref_n)

# --- Load workbook ---
wb = openpyxl.load_workbook('cer.xlsx', data_only=True)
ws = wb['cer']

# --- Add output column headers ---
ws.cell(row=1, column=6,  value="VLM CER")
ws.cell(row=1, column=7,  value="OCR CER")
ws.cell(row=1, column=8,  value="VLM Edit Dist")
ws.cell(row=1, column=9,  value="OCR Edit Dist")
ws.cell(row=1, column=10, value="Ref Len (chars)")

total_vlm_dist = 0
total_ocr_dist = 0
total_ref_len = 0
vlm_cers = []
ocr_cers = []

# --- Compute per-row CER ---
for r in range(2, ws.max_row + 1):
    page  = ws.cell(row=r, column=1).value
    block = ws.cell(row=r, column=2).value
    gt    = ws.cell(row=r, column=3).value   # Ground Truth
    vlm   = ws.cell(row=r, column=4).value   # VLM text
    ocr   = ws.cell(row=r, column=5).value   # OCR text

    if gt is None or normalize(gt) == "":
        continue  # skip rows with no ground truth

    vlm_res = cer(vlm, gt)
    ocr_res = cer(ocr, gt)
    if vlm_res is None or ocr_res is None:
        continue

    vlm_cer, vlm_d, ref_len = vlm_res
    ocr_cer, ocr_d, _       = ocr_res

    ws.cell(row=r, column=6,  value=round(vlm_cer, 4))
    ws.cell(row=r, column=7,  value=round(ocr_cer, 4))
    ws.cell(row=r, column=8,  value=vlm_d)
    ws.cell(row=r, column=9,  value=ocr_d)
    ws.cell(row=r, column=10, value=ref_len)

    total_vlm_dist += vlm_d
    total_ocr_dist += ocr_d
    total_ref_len  += ref_len
    vlm_cers.append(vlm_cer)
    ocr_cers.append(ocr_cer)

# --- Aggregate metrics ---
n = len(vlm_cers)
corpus_vlm_cer = total_vlm_dist / total_ref_len   # micro-avg, char-weighted
corpus_ocr_cer = total_ocr_dist / total_ref_len
macro_vlm_cer  = sum(vlm_cers) / n                # mean of per-block CER
macro_ocr_cer  = sum(ocr_cers) / n

# --- Write summary block below the data ---
summary_start = ws.max_row + 3
ws.cell(row=summary_start,   column=1, value="SUMMARY")
ws.cell(row=summary_start+1, column=1, value="N blocks")
ws.cell(row=summary_start+1, column=2, value=n)
ws.cell(row=summary_start+2, column=1, value="Corpus-level CER (micro-avg, char-weighted)")
ws.cell(row=summary_start+2, column=2, value="VLM")
ws.cell(row=summary_start+2, column=3, value=round(corpus_vlm_cer, 4))
ws.cell(row=summary_start+2, column=4, value="OCR")
ws.cell(row=summary_start+2, column=5, value=round(corpus_ocr_cer, 4))
ws.cell(row=summary_start+3, column=1, value="Macro-average CER (mean of per-block CER)")
ws.cell(row=summary_start+3, column=2, value="VLM")
ws.cell(row=summary_start+3, column=3, value=round(macro_vlm_cer, 4))
ws.cell(row=summary_start+3, column=4, value="OCR")
ws.cell(row=summary_start+3, column=5, value=round(macro_ocr_cer, 4))

wb.save('cer_results.xlsx')

print("N blocks:", n)
print(f"Corpus-level CER  -> VLM: {corpus_vlm_cer:.4f} ({corpus_vlm_cer*100:.2f}%)   "
      f"OCR: {corpus_ocr_cer:.4f} ({corpus_ocr_cer*100:.2f}%)")
print(f"Macro-average CER -> VLM: {macro_vlm_cer:.4f} ({macro_vlm_cer*100:.2f}%)   "
      f"OCR: {macro_ocr_cer:.4f} ({macro_ocr_cer*100:.2f}%)")